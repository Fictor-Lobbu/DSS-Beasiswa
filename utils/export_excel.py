from openpyxl import Workbook

def export_to_excel(df, bobot, kriteria, bo=None, ow=None, filename="hasil_dss.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Perhitungan DSS"

    row = 1

    # =====================
    # JUDUL
    # =====================
    ws[f"A{row}"] = "PERHITUNGAN DSS (BWM + MOORA)"
    row += 2

    # =====================
    # BWM SECTION
    # =====================
    ws[f"A{row}"] = "BWM - Pembobotan"
    row += 1

    ws.append(["Kriteria", "BO", "OW", "Bobot"])

    for i, k in enumerate(kriteria):
        ws.append([
            k,
            bo[i] if bo else "-",
            ow[i] if ow else "-",
            bobot[i]
        ])

    row += len(kriteria) + 2

    # =====================
    # DATA SECTION
    # =====================
    ws[f"A{row}"] = "Data & Transformasi"
    row += 1

    headers = ["Nama"] + kriteria
    ws.append(headers)

    start_data_row = row + 1

    for _, r in df.iterrows():
        ws.append([r["nama"]] + [r[k] for k in kriteria])

    n = len(df)

    row += n + 2

    # =====================
    # NORMALISASI
    # =====================
    ws[f"A{row}"] = "Normalisasi"
    row += 1

    ws.append(headers)

    for i in range(start_data_row, start_data_row + n):
        excel_row = i
        new_row = row + (i - start_data_row) + 1

        ws[f"A{new_row}"] = f"=A{excel_row}"

        for j in range(len(kriteria)):
            col = chr(66 + j)
            formula = f"={col}{excel_row}/SQRT(SUMSQ({col}{start_data_row}:{col}{start_data_row+n-1}))"
            ws[f"{col}{new_row}"] = formula

    row += n + 2

    # =====================
    # TERBOBOT
    # =====================
    ws[f"A{row}"] = "Pembobotan"
    row += 1

    ws.append(headers)

    bobot_start = 5  # lokasi bobot (atas)

    for i in range(n):
        r_excel = row + i + 1
        norm_row = start_data_row + n + 3 + i

        ws[f"A{r_excel}"] = f"=A{norm_row}"

        for j in range(len(kriteria)):
            col = chr(66 + j)
            bobot_cell = f"D{j+5}"  # bobot di tabel atas
            ws[f"{col}{r_excel}"] = f"={col}{norm_row}*{bobot_cell}"

    row += n + 2

    # =====================
    # NILAI Yi
    # =====================
    ws[f"A{row}"] = "Nilai Yi"
    row += 1

    ws.append(["Nama", "Skor"])

    last_col = chr(65 + len(kriteria))

    for i in range(n):
        r_excel = row + i + 1
        terbobot_row = row - n - 1 + i

        ws[f"A{r_excel}"] = f"=A{terbobot_row}"
        ws[f"B{r_excel}"] = f"=SUM(B{terbobot_row}:{last_col}{terbobot_row})"

    wb.save(filename)